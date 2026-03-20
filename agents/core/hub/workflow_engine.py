"""agents/core/hub/workflow_engine.py — HubWorkflowEngine, relocated from
blueprints/agents_hub_bp.py as part of the de-monolith (Phase 3 Slice 1).

Includes the SQL-injection fix for the db_write node
(_validate_write_identifiers): table/column/where-clause identifiers are now
allowlisted against the real target-table schema (via
DatabaseConnectionManager.get_table_schema) before any SQL is built, instead
of being f-string-interpolated with no validation.
"""
import json
import os
import re
import uuid
import agents_hub_bp as _hub_bp  # noqa: F401 — lazy attribute access, see module docstring
import auth  # noqa: F401
from db_exec import run_query as _ss_exec
from core.hub.orchestrator import HubOrchestrator


class HubWorkflowEngine:
    """Hub-aware workflow engine: SQL-Server-backed agents plus extra node types.

    Wraps a base ``core.workflows.engine.WorkflowEngine`` instance by
    composition (kept as ``self._engine`` for its variable-substitution,
    condition-evaluation, and topological-sort helpers) but reimplements node
    execution and agent loading entirely. Adds node types not present in the
    base engine: ``db_read``/``db_write``, ``file_read``/``file_write``,
    ``email_read``/``email_send``, ``loop_start``/``loop_end``, and
    ``approval`` (human-in-the-loop pause/resume).
    """

    def __init__(self, api_key: str, user: dict, run_id: str,
                 workflow_id: str, workflow_name: str):
        """Args:
            api_key: OpenAI API key forwarded to agent-node orchestrators.
            user: The user running this workflow (used for approval/audit records).
            run_id: Unique ID for this workflow run (used in approval context).
            workflow_id: ID of the workflow being executed.
            workflow_name: Display name of the workflow being executed.
        """
        from core.workflows.engine import WorkflowEngine
        self._engine        = WorkflowEngine(api_key)   # kept for _topo_sort + helpers
        self._user          = user
        self._run_id        = run_id
        self._workflow_id   = workflow_id
        self._workflow_name = workflow_name

    # ── Agent loader ──────────────────────────────────────────────────────────

    def _get_agent(self, agent_id: str):
        """Load an agent record (with parsed tools/env_vars) from SQL Server.

        Args:
            agent_id: Agent ID to look up.

        Returns:
            Agent dict with ``tools`` and ``env_vars`` parsed from JSON, or
            None if not found.
        """
        row = _ss_exec('SELECT * FROM hub_agents WHERE id=?', (agent_id,), fetchone=True)
        if not row:
            return None
        row = _hub_bp._fix_row(row)
        row['tools']    = json.loads(row.pop('tools_json', '[]'))
        row['env_vars'] = json.loads(row.pop('env_vars_json', '{}') or '{}')
        return row

    # ── Variable / condition helpers (delegate to WorkflowEngine) ─────────────

    def _substitute_vars(self, template, context):
        """Delegate to ``WorkflowEngine._substitute_vars`` (see that method)."""
        return self._engine._substitute_vars(template, context)

    def _eval_condition(self, expression, context):
        """Delegate to ``WorkflowEngine._eval_condition`` (see that method)."""
        return self._engine._eval_condition(expression, context)

    def _should_execute_node(self, nid, context, nodes, edges):
        """Delegate to ``WorkflowEngine._should_execute_node`` (see that method)."""
        return self._engine._should_execute_node(nid, context, nodes, edges)

    # ── DB query / read ───────────────────────────────────────────────────────

    def _exec_query_db(self, node, context):
        """Delegate to ``WorkflowEngine._exec_query_db`` (see that method)."""
        return self._engine._exec_query_db(node, context)

    def _exec_db_read(self, node, context):
        """Run a ``db_read`` node — identical implementation to ``query_db``."""
        return self._engine._exec_query_db(node, context)   # same as query_db

    # ── HTTP request ─────────────────────────────────────────────────────────

    def _exec_http_request(self, node, context):
        """Delegate to ``WorkflowEngine._exec_http_request`` (see that method)."""
        return self._engine._exec_http_request(node, context)

    # ── DB write ──────────────────────────────────────────────────────────────

    # SQL keywords that legitimately appear in a where_clause but aren't
    # real column identifiers — excluded from the allowlist check below.
    _WHERE_CLAUSE_KEYWORDS = {
        'and', 'or', 'not', 'null', 'is', 'in', 'like', 'between',
        'true', 'false', 'exists', 'select', 'as', 'on',
    }
    _IDENTIFIER_SHAPE_RE = re.compile(r'^[A-Za-z_][A-Za-z0-9_.\[\]]*$')

    def _validate_write_identifiers(self, cfg, table, cols, where):
        """Allowlist table/column/where-clause identifiers for a db_write
        node against the target table's real schema, before any SQL is
        built. Returns None if everything checks out, else an error string.

        Values are already safely bound via '?' placeholders elsewhere in
        _exec_db_write — this only guards the identifier positions (table
        name, column names, and any bare identifiers inside a free-form
        where_clause), which previously went straight into the SQL string
        via f-string interpolation with no validation at all.
        """
        if not self._IDENTIFIER_SHAPE_RE.match(table or ''):
            return f"Invalid table name: {table!r}"

        try:
            from database_manager import DatabaseConnectionManager
            schema = DatabaseConnectionManager().get_table_schema(cfg, table)
        except Exception as e:
            return f"Could not verify table schema for {table!r}: {e}"

        real_cols = {c['name'].lower() for c in (schema or {}).get('columns', [])}
        if not real_cols:
            return f"Unknown table: {table!r}"

        for col in cols:
            if col.lower() not in real_cols:
                return f"Unknown column: {col!r}"

        if where:
            for tok in re.findall(r'[A-Za-z_][A-Za-z0-9_]*', where):
                low = tok.lower()
                if low in self._WHERE_CLAUSE_KEYWORDS:
                    continue
                if low not in real_cols:
                    return f"Invalid identifier in where clause: {tok!r}"

        return None

    def _exec_db_write(self, node, context):
        """Execute an INSERT/UPDATE/UPSERT against a named database connection.

        Resolves the connection by ``node['connection_name']`` (falling back
        to the first configured connection), builds the SQL statement from
        ``node['field_mapping']`` (column -> value-template pairs, with
        ``{{var}}`` substitution), and dispatches to the matching driver
        (pyodbc/psycopg2/pymysql/sqlite3). UPSERT uses a SQL Server MERGE
        statement when the connection type is SQL Server/MSSQL, otherwise
        falls back to a plain INSERT.

        Args:
            node: Workflow node dict with ``connection_name``, ``table_name``,
                ``operation`` ("INSERT"/"UPDATE"/"UPSERT"), ``field_mapping``,
                and ``where_clause`` (for UPDATE/UPSERT).
            context: Run context used for variable substitution.

        Returns:
            ``{"rows_affected": N, "operation": ..., "table": ...}`` on
            success, or ``{"error": "..."}`` on failure.
        """
        try:
            from database_manager import DatabaseConnectionManager
            mgr  = DatabaseConnectionManager()
            conn_name = node.get('connection_name', '')
            conns     = mgr.load_connections()
            cfg       = next((c for c in conns if c['name'] == conn_name), None)
            if not cfg and conns:
                cfg = conns[0]
            if not cfg:
                return {"error": "No database connection configured"}

            table     = self._substitute_vars(node.get('table_name', ''), context)
            operation = node.get('operation', 'INSERT').upper()
            mapping   = node.get('field_mapping', [])
            where     = self._substitute_vars(node.get('where_clause', ''), context)

            if not table:
                return {"error": "Table name is required"}

            # Build column/value pairs
            cols, vals = [], []
            for m in mapping:
                col = m.get('column', '').strip()
                val = self._substitute_vars(m.get('value', ''), context)
                if col:
                    cols.append(col)
                    vals.append(val)

            if not cols:
                return {"error": "No field mapping configured"}

            id_error = self._validate_write_identifiers(cfg, table, cols, where)
            if id_error:
                return {"error": id_error}

            db_type = cfg.get('type', '').lower()
            server   = cfg.get('server', '')
            database = cfg.get('database', '')
            username = cfg.get('username', '')
            password = cfg.get('password', '')
            port     = cfg.get('port', '')

            if operation == 'INSERT':
                sql = f"INSERT INTO {table} ({', '.join(cols)}) VALUES ({', '.join(['?' for _ in cols])})"
            elif operation == 'UPDATE':
                set_clause = ', '.join([f"{c}=?" for c in cols])
                sql = f"UPDATE {table} SET {set_clause} WHERE {where or '1=1'}"
            elif operation == 'UPSERT':
                # Basic MERGE for SQL Server; fallback to INSERT for others
                if any(t in db_type for t in ('sqlserver', 'mssql')):
                    set_clause   = ', '.join([f"T.{c}=S.{c}" for c in cols])
                    s_cols       = ', '.join(cols)
                    s_vals       = ', '.join([f'? AS {c}' for c in cols])
                    sql = (f"MERGE INTO {table} AS T "
                           f"USING (SELECT {s_vals}) AS S ON ({where or '1=0'}) "
                           f"WHEN MATCHED THEN UPDATE SET {set_clause} "
                           f"WHEN NOT MATCHED THEN INSERT ({s_cols}) VALUES ({', '.join(['S.'+c for c in cols])});")
                else:
                    sql = f"INSERT INTO {table} ({', '.join(cols)}) VALUES ({', '.join(['?' for _ in cols])})"
            else:
                return {"error": f"Unknown operation: {operation}"}

            rows_affected = 0
            if any(t in db_type for t in ('sqlserver', 'mssql', 'sql server')):
                import pyodbc
                cs = (f"DRIVER={{ODBC Driver 17 for SQL Server}};"
                      f"SERVER={server};DATABASE={database};"
                      f"UID={username};PWD={password};")
                with pyodbc.connect(cs, timeout=30) as c:
                    cur = c.cursor(); cur.execute(sql, vals); c.commit()
                    rows_affected = cur.rowcount
            elif any(t in db_type for t in ('postgres', 'postgresql')):
                import psycopg2
                with psycopg2.connect(host=server, dbname=database, user=username,
                                      password=password, port=int(port or 5432)) as c:
                    cur = c.cursor(); cur.execute(sql, vals); c.commit()
                    rows_affected = cur.rowcount
            elif 'mysql' in db_type:
                import pymysql
                with pymysql.connect(host=server, database=database, user=username,
                                     password=password, port=int(port or 3306)) as c:
                    cur = c.cursor(); cur.execute(sql, vals); c.commit()
                    rows_affected = cur.rowcount
            elif 'sqlite' in db_type:
                import sqlite3
                with sqlite3.connect(database) as c:
                    cur = c.cursor(); cur.execute(sql, vals); c.commit()
                    rows_affected = cur.rowcount
            else:
                return {"error": f"Unsupported db type: {db_type}"}

            return {"rows_affected": rows_affected, "operation": operation, "table": table}
        except Exception as e:
            return {"error": str(e)}

    # ── File read ─────────────────────────────────────────────────────────────

    def _exec_file_read(self, node, context):
        """Read rows from an Excel, CSV, or JSON file.

        Args:
            node: Workflow node dict with ``file_path``, ``file_format``
                ("excel"/"csv"/"json"), and ``sheet_name`` (Excel only).
            context: Run context used for variable substitution in ``file_path``.

        Returns:
            ``{"rows": [...], "count": N}`` (rows as dicts keyed by header/
            column name) on success, or ``{"error": "...", "rows": []}``.
        """
        try:
            file_path  = self._substitute_vars(node.get('file_path', ''), context)
            file_fmt   = node.get('file_format', 'excel').lower()
            sheet_name = node.get('sheet_name', '').strip() or None

            if not file_path:
                return {"error": "File path is required", "rows": []}

            if file_fmt == 'excel':
                import openpyxl
                wb   = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                ws   = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
                rows_iter = list(ws.rows)
                if not rows_iter:
                    return {"rows": [], "count": 0}
                headers = [str(c.value or f'col{i}') for i, c in enumerate(rows_iter[0])]
                rows    = []
                for row in rows_iter[1:]:
                    rows.append({headers[i]: (str(c.value) if c.value is not None else '')
                                 for i, c in enumerate(row) if i < len(headers)})
                return {"rows": rows, "count": len(rows)}

            elif file_fmt == 'csv':
                import csv
                rows = []
                with open(file_path, newline='', encoding='utf-8-sig') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        rows.append(dict(row))
                return {"rows": rows, "count": len(rows)}

            elif file_fmt == 'json':
                import json as _json
                with open(file_path, 'r', encoding='utf-8') as f:
                    data = _json.load(f)
                if isinstance(data, list):
                    return {"rows": data, "count": len(data)}
                return {"rows": [data], "count": 1}

            return {"error": f"Unknown format: {file_fmt}", "rows": []}
        except Exception as e:
            return {"error": str(e), "rows": []}

    # ── File write ────────────────────────────────────────────────────────────

    def _exec_file_write(self, node, context):
        """Write rows (from a context variable or last_output) to an Excel/CSV file.

        Rows are read from ``context['vars'][node['data_var']]`` (or
        ``context['last_output']`` if ``data_var`` is unset), parsed as JSON
        if they're a string. Column headers/values come from
        ``node['column_mapping']`` (with per-row ``{{loop_item}}``
        substitution) when provided, otherwise inferred from dict keys or a
        single ``value`` column.

        Args:
            node: Workflow node dict with ``file_path``, ``file_format``
                ("excel"/"csv"), ``sheet_name``, ``data_var``, ``column_mapping``.
            context: Run context supplying the data and variable substitution.

        Returns:
            ``{"written": N, "file": path, "format": fmt}`` on success, or
            ``{"error": "..."}`` on failure.
        """
        try:
            file_path  = self._substitute_vars(node.get('file_path', ''), context)
            file_fmt   = node.get('file_format', 'excel').lower()
            sheet_name = node.get('sheet_name', 'Sheet1').strip() or 'Sheet1'
            data_var   = node.get('data_var', '').strip()
            col_map    = node.get('column_mapping', [])

            if not file_path:
                return {"error": "Output file path is required"}

            raw = context['vars'].get(data_var, context.get('last_output', '[]')) if data_var else context.get('last_output', '[]')
            try:
                rows = json.loads(raw) if isinstance(raw, str) else raw
                if not isinstance(rows, list):
                    rows = [rows]
            except Exception:
                rows = [{'value': str(raw)}]

            # Build headers and extract values
            if col_map and any(m.get('column') for m in col_map):
                headers = [m['column'] for m in col_map if m.get('column')]
                def get_row(item):
                    item_ctx = {**context, 'vars': {**context['vars'],
                                'loop_item': json.dumps(item) if isinstance(item, dict) else str(item)}}
                    return [self._substitute_vars(m.get('field', ''), item_ctx)
                            for m in col_map if m.get('column')]
            else:
                if rows and isinstance(rows[0], dict):
                    headers  = list(rows[0].keys())
                    get_row  = lambda item: [str(item.get(h, '')) for h in headers]
                else:
                    headers  = ['value']
                    get_row  = lambda item: [str(item)]

            import os
            os.makedirs(os.path.dirname(os.path.abspath(file_path)), exist_ok=True)

            if file_fmt == 'excel':
                import openpyxl
                wb = openpyxl.Workbook()
                ws = wb.active; ws.title = sheet_name
                ws.append(headers)
                for item in rows:
                    ws.append(get_row(item))
                wb.save(file_path)

            elif file_fmt == 'csv':
                import csv
                with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
                    writer = csv.writer(f)
                    writer.writerow(headers)
                    for item in rows:
                        writer.writerow(get_row(item))

            return {"written": len(rows), "file": file_path, "format": file_fmt}
        except Exception as e:
            return {"error": str(e)}

    # ── Email read (IMAP) ─────────────────────────────────────────────────────

    def _exec_email_read(self, node, context):
        """Fetch recent emails from an IMAP mailbox.

        Args:
            node: Workflow node dict with ``imap_host``, ``imap_port``,
                ``imap_user``, ``imap_pass``, ``folder``, ``filter_unread``,
                ``max_emails``.
            context: Run context (unused for substitution here; IMAP
                settings are taken from the node directly).

        Returns:
            ``{"emails": [...], "count": N}`` where each email dict has
            ``from``, ``to``, ``subject``, ``date``, ``body`` (truncated to
            10000 chars), ``message_id`` — or ``{"error": "...", "emails": []}``.
        """
        try:
            import imaplib, email as _email_lib, ssl
            from email.header import decode_header

            host       = node.get('imap_host', '')
            port       = int(node.get('imap_port', 993) or 993)
            user       = node.get('imap_user', '')
            password   = node.get('imap_pass', '')
            folder     = node.get('folder', 'INBOX').strip() or 'INBOX'
            unread     = node.get('filter_unread', True)
            max_emails = int(node.get('max_emails', 10) or 10)

            if not host or not user or not password:
                return {"error": "IMAP host, user, and password are required", "emails": []}

            ctx  = ssl.create_default_context()
            mail = imaplib.IMAP4_SSL(host, port, ssl_context=ctx)
            mail.login(user, password)
            mail.select(folder)

            status, uids = mail.search(None, 'UNSEEN' if unread else 'ALL')
            uid_list = uids[0].split()[-max_emails:]   # take latest N

            emails = []
            for uid in uid_list:
                _, msg_data = mail.fetch(uid, '(RFC822)')
                msg = _email_lib.message_from_bytes(msg_data[0][1])

                def decode_str(s):
                    parts = decode_header(s or '')
                    return ''.join(
                        p.decode(enc or 'utf-8') if isinstance(p, bytes) else p
                        for p, enc in parts)

                body = ''
                if msg.is_multipart():
                    for part in msg.walk():
                        ct = part.get_content_type()
                        if ct in ('text/plain', 'text/html') and not part.get('Content-Disposition'):
                            charset = part.get_content_charset() or 'utf-8'
                            body = part.get_payload(decode=True).decode(charset, errors='replace')
                            if ct == 'text/plain':
                                break
                else:
                    charset = msg.get_content_charset() or 'utf-8'
                    body = msg.get_payload(decode=True).decode(charset, errors='replace')

                emails.append({
                    'from':       decode_str(msg.get('From', '')),
                    'to':         decode_str(msg.get('To', '')),
                    'subject':    decode_str(msg.get('Subject', '')),
                    'date':       msg.get('Date', ''),
                    'body':       body[:10000],
                    'message_id': msg.get('Message-ID', ''),
                })

            mail.close(); mail.logout()
            return {"emails": emails, "count": len(emails)}
        except Exception as e:
            return {"error": str(e), "emails": []}

    # ── Email send (SMTP) ─────────────────────────────────────────────────────

    def _exec_email_send(self, node, context):
        """Send an email via SMTP using server settings from environment variables.

        SMTP connection details (host/port/user/pass/from) come from the
        ``SMTP_HOST``/``SMTP_PORT``/``SMTP_USER``/``SMTP_PASS``/``SMTP_FROM``
        env vars, not from the node — only recipients/subject/body are
        node-configured. Uses SSL directly on port 465, otherwise STARTTLS.

        Args:
            node: Workflow node dict with ``to``, ``cc``, ``subject``, ``body``,
                ``is_html``.
            context: Run context used for variable substitution.

        Returns:
            ``{"sent": True, "to": ..., "subject": ...}`` on success, or
            ``{"error": "...", "sent": False}`` / ``{"error": "..."}`` on failure.
        """
        try:
            import smtplib, ssl
            from email.mime.multipart import MIMEMultipart
            from email.mime.text      import MIMEText

            to_addr  = self._substitute_vars(node.get('to', ''), context).strip()
            cc_addr  = self._substitute_vars(node.get('cc', ''), context).strip()
            subject  = self._substitute_vars(node.get('subject', ''), context)
            body     = self._substitute_vars(node.get('body', ''), context)
            is_html  = node.get('is_html', False)

            if not to_addr:
                return {"error": "Recipient (To) is required"}

            smtp_host = os.environ.get('SMTP_HOST', '')
            smtp_port = int(os.environ.get('SMTP_PORT', 587) or 587)
            smtp_user = os.environ.get('SMTP_USER', '')
            smtp_pass = os.environ.get('SMTP_PASS', '')
            smtp_from = os.environ.get('SMTP_FROM', smtp_user)

            if not smtp_host or not smtp_user:
                return {"error": "SMTP not configured. Set SMTP_HOST, SMTP_PORT, SMTP_USER, SMTP_PASS, SMTP_FROM env vars."}

            msg = MIMEMultipart('alternative')
            msg['From']    = smtp_from
            msg['To']      = to_addr
            if cc_addr:
                msg['Cc']  = cc_addr
            msg['Subject'] = subject

            ct = 'html' if is_html else 'plain'
            msg.attach(MIMEText(body, ct, 'utf-8'))

            all_recipients = [a.strip() for a in (to_addr + ',' + cc_addr).split(',') if a.strip()]

            ctx = ssl.create_default_context()
            if smtp_port == 465:
                with smtplib.SMTP_SSL(smtp_host, smtp_port, context=ctx) as server:
                    server.login(smtp_user, smtp_pass)
                    server.sendmail(smtp_from, all_recipients, msg.as_string())
            else:
                with smtplib.SMTP(smtp_host, smtp_port) as server:
                    server.starttls(context=ctx)
                    server.login(smtp_user, smtp_pass)
                    server.sendmail(smtp_from, all_recipients, msg.as_string())

            return {"sent": True, "to": to_addr, "subject": subject}
        except Exception as e:
            return {"error": str(e), "sent": False}

    # ── Loop body finder ──────────────────────────────────────────────────────

    def _find_loop_body(self, loop_start_id, nodes, edges):
        """Return (body_node_ids_list, loop_end_id) reachable via loop_body port."""
        body, visited, loop_end_id = [], set(), None
        queue = [tgt for e in edges
                 if (e.get('source') or e.get('from')) == loop_start_id
                 and e.get('source_port') == 'loop_body'
                 for tgt in [(e.get('target') or e.get('to'))] if tgt]
        while queue:
            nid = queue.pop(0)
            if nid in visited:
                continue
            visited.add(nid)
            node = next((n for n in nodes if n['id'] == nid), None)
            if not node:
                continue
            if node.get('type') == 'loop_end':
                loop_end_id = nid
            else:
                body.append(nid)
                for e in edges:
                    src = e.get('source') or e.get('from')
                    tgt = e.get('target') or e.get('to')
                    if src == nid and tgt not in visited:
                        queue.append(tgt)
        return body, loop_end_id

    # ── Entry point ───────────────────────────────────────────────────────────

    def execute(self, workflow: dict, input_data: str):
        """Run a workflow from the start, streaming NDJSON progress events.

        Args:
            workflow: Workflow dict with ``nodes``, ``edges``, ``name``, and
                ``execution_mode`` ("sequential" or "parallel").
            input_data: The trigger input value fed to the first node(s).

        Yields:
            JSON-encoded (NDJSON) strings for each lifecycle/node event,
            starting with a ``wf_start`` event. May end in ``wf_complete`` or,
            if an ``approval`` node is hit, ``wf_paused``.
        """
        nodes = workflow.get('nodes', [])
        edges = workflow.get('edges', [])
        mode  = workflow.get('execution_mode', 'sequential')
        yield json.dumps({'type': 'wf_start', 'workflow': workflow.get('name', ''),
                          'nodes': len(nodes)}) + '\n'
        if mode == 'parallel':
            yield from self._run_parallel(nodes, input_data)
        else:
            yield from self._run_sequential(nodes, edges, input_data)

    # ── Sequential execution ──────────────────────────────────────────────────

    def _run_sequential(self, nodes, edges, input_data):
        """Build the initial run context and execute nodes in topological order.

        Args:
            nodes: All workflow nodes.
            edges: All workflow edges.
            input_data: Trigger input value for the first node.

        Yields:
            NDJSON event strings from ``_execute_nodes``.
        """
        order   = self._engine._topo_sort(nodes, edges)
        context = {
            'last_output':       input_data,
            'outputs':           {},
            'vars':              {'trigger_input': input_data},
            'condition_results': {},
        }
        yield from self._execute_nodes(nodes, edges, order, context)

    # ── Core node executor (shared by first run and resume) ───────────────────

    @staticmethod
    def _all_parents_skipped(nid, edges, skipped):
        """True when every parent edge comes from a node in the skipped set.
        Nodes with no parents (start nodes) are never considered all-skipped."""
        parents = [e.get('source') or e.get('from')
                   for e in edges
                   if (e.get('target') or e.get('to')) == nid]
        return bool(parents) and all(p in skipped for p in parents)

    def _execute_nodes(self, nodes, edges, order, context, resume_after=None):
        """Execute nodes in topological order, dispatching by node type.

        This is the shared engine for both a fresh run (``_run_sequential``)
        and a post-approval resume (``resume``). Supports many more node
        types than the base ``WorkflowEngine``: trigger/input, db_read,
        file_read, email_read, loop_start/loop_end, email_send, db_write,
        file_write, approval (pauses and returns), agent, query_db, condition,
        transform, http_request, set_variable, delay, output.

        When resume_after is set (a node ID), all nodes up to and including
        that node are dry-run (condition checks only, no actual execution) so
        that the `skipped` set is correctly populated before real execution
        begins.  This prevents cascade-skip from failing during a resume
        because nodes that were skipped in the first half are never in the
        remaining `order`.

        Args:
            nodes: All workflow nodes.
            edges: All workflow edges.
            order: Node IDs in topological execution order.
            context: Mutable run context (``last_output``, ``outputs``,
                ``vars``, ``condition_results``, and on resume also
                ``total_tokens``/``input_tokens``/``output_tokens``).
            resume_after: If set, the node ID after which to resume real
                execution; everything up to and including it is dry-run only.

        Yields:
            JSON-encoded (NDJSON) per-node events (``node_skipped``,
            ``node_start``, per-type result events, ``node_complete``), then
            either ``wf_complete`` or, if an ``approval`` node is hit,
            ``approval_required``/``wf_paused`` (which also returns early).
        """
        import time as _t

        # Build the correct skipped set before starting real execution
        skipped = set()
        if resume_after and resume_after in order:
            resume_idx = order.index(resume_after) + 1
            for nid in order[:resume_idx]:
                if not self._should_execute_node(nid, context, nodes, edges):
                    skipped.add(nid)
                elif self._all_parents_skipped(nid, edges, skipped):
                    skipped.add(nid)
            order = order[resume_idx:]   # only execute the remainder

        for nid in order:
            node  = next((n for n in nodes if n['id'] == nid), None)
            if not node:
                continue
            ntype = node.get('type', 'agent')

            # 0 — loop body handled: skip nodes that were executed inside _exec loop_start
            if nid in context.get('loop_handled', set()):
                yield json.dumps({'type': 'node_skipped', 'node_id': nid,
                                  'reason': 'in_loop'}) + '\n'
                continue

            # 1 — condition-branch skip (direct conditional edge says "don't take this path")
            if not self._should_execute_node(nid, context, nodes, edges):
                skipped.add(nid)
                yield json.dumps({'type': 'node_skipped', 'node_id': nid}) + '\n'
                continue

            # 2 — cascade skip: all parents were skipped, so there is no live path here
            if self._all_parents_skipped(nid, edges, skipped):
                skipped.add(nid)
                yield json.dumps({'type': 'node_skipped', 'node_id': nid}) + '\n'
                continue

            yield json.dumps({'type': 'node_start', 'node_id': nid,
                              'node_type': ntype, 'name': node.get('label', nid)}) + '\n'

            output_var = node.get('output_var', '').strip()

            # ── trigger ───────────────────────────────────────────────────
            if ntype in ('trigger', 'input'):
                val = context.get('last_output', '')
                if output_var:
                    context['vars'][output_var] = val
                context['outputs'][nid] = val

            # ── db_read (source) ──────────────────────────────────────────
            elif ntype == 'db_read':
                result = self._exec_db_read(node, context)
                val    = json.dumps(result.get('rows', []), ensure_ascii=False)
                context['last_output'] = val
                context['outputs'][nid] = val
                if output_var:
                    context['vars'][output_var] = val
                yield json.dumps({'type': 'query_result', 'node_id': nid,
                                  'count': result.get('count', 0),
                                  'error': result.get('error')}) + '\n'

            # ── file_read (source) ────────────────────────────────────────
            elif ntype == 'file_read':
                result = self._exec_file_read(node, context)
                val    = json.dumps(result.get('rows', []), ensure_ascii=False)
                context['last_output'] = val
                context['outputs'][nid] = val
                if output_var:
                    context['vars'][output_var] = val
                yield json.dumps({'type': 'file_read_result', 'node_id': nid,
                                  'count': result.get('count', 0),
                                  'error': result.get('error')}) + '\n'

            # ── email_read (source) ───────────────────────────────────────
            elif ntype == 'email_read':
                result = self._exec_email_read(node, context)
                val    = json.dumps(result.get('emails', []), ensure_ascii=False)
                context['last_output'] = val
                context['outputs'][nid] = val
                if output_var:
                    context['vars'][output_var] = val
                yield json.dumps({'type': 'email_read_result', 'node_id': nid,
                                  'count': result.get('count', 0),
                                  'error': result.get('error')}) + '\n'

            # ── loop_start ────────────────────────────────────────────────
            elif ntype == 'loop_start':
                loop_var   = node.get('loop_var', '').strip()
                item_var   = node.get('item_var', 'loop_item').strip() or 'loop_item'
                index_var  = node.get('index_var', 'loop_index').strip() or 'loop_index'
                result_var = node.get('result_var', 'loop_results').strip() or 'loop_results'

                raw = context['vars'].get(loop_var, context.get('last_output', '[]')) if loop_var else context.get('last_output', '[]')
                try:
                    items = json.loads(raw) if isinstance(raw, str) else raw
                    if not isinstance(items, list):
                        items = [items]
                except Exception:
                    items = [raw]

                body_ids, loop_end_id = self._find_loop_body(nid, nodes, edges)
                context.setdefault('loop_handled', set())
                context['loop_handled'].update(body_ids)
                if loop_end_id:
                    context['loop_handled'].add(loop_end_id)

                body_order  = [n for n in self._engine._topo_sort(nodes, edges) if n in set(body_ids)]
                body_nodes  = [n for n in nodes if n['id'] in set(body_ids)]

                results = []
                yield json.dumps({'type': 'loop_start', 'node_id': nid,
                                  'total': len(items), 'item_var': item_var}) + '\n'
                for i, item in enumerate(items):
                    item_str = json.dumps(item, ensure_ascii=False) if isinstance(item, dict) else str(item)
                    yield json.dumps({'type': 'loop_iter', 'node_id': nid,
                                      'index': i, 'total': len(items)}) + '\n'
                    child_ctx = {
                        'last_output':       item_str,
                        'outputs':           {},
                        'vars':              {**context['vars'], item_var: item_str, index_var: str(i)},
                        'condition_results': dict(context.get('condition_results', {})),
                    }
                    for chunk in self._execute_nodes(body_nodes, edges, body_order, child_ctx):
                        try:
                            d = json.loads(chunk.strip())
                            d['loop_index'] = i
                            yield json.dumps(d) + '\n'
                        except Exception:
                            yield chunk
                    results.append(child_ctx.get('last_output', ''))

                results_val = json.dumps(results, ensure_ascii=False)
                context['vars'][result_var] = results_val
                if output_var:
                    context['vars'][output_var] = results_val
                context['last_output'] = results_val
                context['outputs'][nid] = results_val
                yield json.dumps({'type': 'loop_complete', 'node_id': nid,
                                  'iterations': len(items)}) + '\n'

            # ── loop_end ──────────────────────────────────────────────────
            elif ntype == 'loop_end':
                pass   # handled via loop_handled skip above

            # ── email_send ────────────────────────────────────────────────
            elif ntype == 'email_send':
                result = self._exec_email_send(node, context)
                val    = json.dumps(result, ensure_ascii=False)
                context['last_output'] = val
                context['outputs'][nid] = val
                if output_var:
                    context['vars'][output_var] = val
                yield json.dumps({'type': 'email_send_result', 'node_id': nid,
                                  'sent':  result.get('sent', False),
                                  'to':    result.get('to', ''),
                                  'error': result.get('error')}) + '\n'

            # ── db_write ──────────────────────────────────────────────────
            elif ntype == 'db_write':
                result = self._exec_db_write(node, context)
                val    = json.dumps(result, ensure_ascii=False)
                context['last_output'] = val
                context['outputs'][nid] = val
                yield json.dumps({'type': 'db_write_result', 'node_id': nid,
                                  'rows_affected': result.get('rows_affected', 0),
                                  'error':         result.get('error')}) + '\n'

            # ── file_write ────────────────────────────────────────────────
            elif ntype == 'file_write':
                result = self._exec_file_write(node, context)
                val    = json.dumps(result, ensure_ascii=False)
                context['last_output'] = val
                context['outputs'][nid] = val
                yield json.dumps({'type': 'file_write_result', 'node_id': nid,
                                  'written': result.get('written', 0),
                                  'file':    result.get('file', ''),
                                  'error':   result.get('error')}) + '\n'

            # ── approval ──────────────────────────────────────────────────
            elif ntype == 'approval':
                approval_id = str(uuid.uuid4())
                title       = node.get('approval_title') or node.get('label') or 'Approval Required'
                ctx_text    = self._substitute_vars(node.get('approval_context', ''), context)
                assigned_to = node.get('assigned_to_user_id')

                # ── Embed full resume context so the workflow can continue after approval
                resume_ctx = {
                    'paused_at_node':   nid,
                    'execution_order':  order,
                    'context': {
                        'vars':              {k: str(v)[:5000] for k, v in context.get('vars', {}).items()},
                        'last_output':       str(context.get('last_output', ''))[:5000],
                        'outputs':           {k: str(v)[:2000] for k, v in context.get('outputs', {}).items()},
                        'condition_results': {k: bool(v) for k, v in context.get('condition_results', {}).items()},
                    },
                }
                ctx_blob = json.dumps({
                    'workflow_id':    self._workflow_id,
                    'workflow_name':  self._workflow_name,
                    'run_id':         self._run_id,
                    'node_id':        nid,
                    'node_label':     node.get('label', ''),
                    'custom_context': ctx_text,
                    'current_output': str(context.get('last_output', ''))[:2000],
                    'requested_by':   self._user.get('username', ''),
                    'resume':         resume_ctx,           # ← resume state saved here
                })
                _ss_exec("""
                    INSERT INTO hub_approvals
                        (approval_id, request_type, workflow_id, workflow_name,
                         run_id, node_id, requested_by_user_id,
                         assigned_to_user_id, title, context_json)
                    VALUES (?, 'workflow', ?, ?, ?, ?, ?, ?, ?, ?)
                """, (approval_id, self._workflow_id, self._workflow_name,
                      self._run_id, nid, self._user.get('id'),
                      int(assigned_to) if assigned_to else None, title, ctx_blob))

                yield json.dumps({
                    'type': 'approval_required', 'node_id': nid,
                    'approval_id': approval_id, 'title': title,
                    'workflow_id': self._workflow_id,
                    'assigned_to_user_id': assigned_to,
                }) + '\n'
                yield json.dumps({
                    'type': 'wf_paused', 'approval_id': approval_id,
                    'workflow_id': self._workflow_id,
                    'final_output': context.get('last_output', ''),
                }) + '\n'
                return

            # ── agent ─────────────────────────────────────────────────────
            elif ntype == 'agent' and node.get('agent_id'):
                agent = self._get_agent(node['agent_id'])
                if agent:
                    if _hub_bp._APPROVAL_ENABLED:
                        _hub_bp._ensure_approval_tool()
                    tool_configs = _hub_bp._extract_tool_configs(agent.get('tools', []))
                    if _hub_bp._APPROVAL_ENABLED and 'request_human_approval' not in _hub_bp._tool_names(agent['tools']):
                        agent = {**agent, 'tools': agent['tools'] + ['request_human_approval']}
                    hub_ctx = {
                        'user':            self._user,
                        'agent_id':        node['agent_id'],
                        'agent_name':      agent.get('name', ''),
                        'convo_id':        self._run_id,
                        'agent_env_vars':  agent.get('env_vars', {}),
                    }
                    question = self._substitute_vars(
                        node.get('question', '') or context['last_output'], context)
                    orch = HubOrchestrator(self._engine.api_key, hub_ctx, tool_configs)
                    for chunk in orch.run(question, agent):
                        try:
                            d = json.loads(chunk.strip())
                            d['node_id'] = nid
                            if d.get('type') == 'final':
                                val = d.get('content', '')
                                context['last_output']  = val
                                context['outputs'][nid] = val
                                if output_var:
                                    context['vars'][output_var] = val
                                node_tokens_used   = d.get('tokens_used')   or 0
                                node_input_tokens  = d.get('input_tokens')  or 0
                                node_output_tokens = d.get('output_tokens') or 0
                                context['total_tokens']  = context.get('total_tokens', 0)  + node_tokens_used
                                context['input_tokens']  = context.get('input_tokens', 0)  + node_input_tokens
                                context['output_tokens'] = context.get('output_tokens', 0) + node_output_tokens
                                # Record per-node so cost is priced against the
                                # model that node actually ran on — a workflow
                                # can mix agents on different models, so a
                                # single aggregate row at the end of the run
                                # can't be priced accurately.
                                if self._user and self._user.get('id') and (node_tokens_used or node_input_tokens or node_output_tokens):
                                    try:
                                        import token_limits as _tl_node
                                        _tl_node.record_tokens(
                                            self._user, call_type='hub_workflow',
                                            agent_name=agent.get('name', ''),
                                            question=question,
                                            actual_tokens=node_tokens_used or None,
                                            response=val,
                                            input_tokens=node_input_tokens,
                                            output_tokens=node_output_tokens,
                                            model=agent.get('model') or 'gpt-4o',
                                            agent_id=node['agent_id'],
                                        )
                                    except Exception:
                                        pass
                            elif (d.get('type') == 'tool_result'
                                  and d.get('tool') == 'request_human_approval'):
                                inner = (d.get('result') or {}).get('result') or {}
                                aid   = inner.get('approval_id') if isinstance(inner, dict) else None
                                if aid:
                                    yield json.dumps({'type': 'approval_requested',
                                                      'approval_id': aid, 'node_id': nid}) + '\n'
                            yield json.dumps(d) + '\n'
                        except Exception:
                            yield chunk

            # ── query_db ──────────────────────────────────────────────────
            elif ntype == 'query_db':
                result = self._exec_query_db(node, context)
                val    = json.dumps(result, ensure_ascii=False)
                context['last_output'] = val
                context['outputs'][nid] = val
                if output_var:
                    context['vars'][output_var] = val
                yield json.dumps({'type': 'query_result', 'node_id': nid,
                                  'rows':  result.get('rows', [])[:50],
                                  'count': result.get('count', 0),
                                  'error': result.get('error')}) + '\n'

            # ── condition ─────────────────────────────────────────────────
            elif ntype == 'condition':
                expression = node.get('expression', node.get('condition', ''))
                result     = self._eval_condition(expression, context)
                context['condition_results'][nid] = result
                yield json.dumps({'type': 'condition_result', 'node_id': nid,
                                  'result': result}) + '\n'

            # ── transform ─────────────────────────────────────────────────
            elif ntype == 'transform':
                val = self._substitute_vars(
                    node.get('expression', '{{trigger_input}}'), context)
                context['last_output'] = val
                context['outputs'][nid] = val
                if output_var:
                    context['vars'][output_var] = val

            # ── http_request ──────────────────────────────────────────────
            elif ntype == 'http_request':
                result = self._exec_http_request(node, context)
                val    = json.dumps(result, ensure_ascii=False)
                context['last_output'] = val
                context['outputs'][nid] = val
                if output_var:
                    context['vars'][output_var] = val
                yield json.dumps({'type': 'http_result', 'node_id': nid,
                                  'status': result.get('status'),
                                  'error':  result.get('error')}) + '\n'

            # ── set_variable ──────────────────────────────────────────────
            elif ntype == 'set_variable':
                var_name = node.get('var_name', '').strip()
                val      = self._substitute_vars(node.get('value', ''), context)
                if var_name:
                    context['vars'][var_name] = val
                context['last_output'] = val
                context['outputs'][nid] = val

            # ── delay ─────────────────────────────────────────────────────
            elif ntype == 'delay':
                _t.sleep(min(node.get('delay_seconds', 1), 5))

            # ── output ────────────────────────────────────────────────────
            elif ntype == 'output':
                dv  = node.get('display_var', '').strip()
                val = context['vars'].get(dv, context['last_output']) if dv else context['last_output']
                context['last_output'] = val
                context['outputs'][nid] = val

            yield json.dumps({'type': 'node_complete', 'node_id': nid,
                              'output_preview': str(context['outputs'].get(nid, ''))[:300]}) + '\n'

        yield json.dumps({
            'type':          'wf_complete',
            'final_output':  context.get('last_output', ''),
            'vars':          {k: str(v)[:500] for k, v in context.get('vars', {}).items()},
            'total_tokens':  context.get('total_tokens',  0),
            'input_tokens':  context.get('input_tokens',  0),
            'output_tokens': context.get('output_tokens', 0),
        }) + '\n'

    # ── Resume after approval ─────────────────────────────────────────────────

    def resume(self, workflow, approval_id: str, approved: bool):
        """Continue a workflow that was paused at an approval node.

        Restores the run context that was saved into ``hub_approvals.context_json``
        (see the ``approval`` node branch in ``_execute_nodes``) and, if approved,
        dry-runs all pre-pause nodes to rebuild the cascade-skip set before
        resuming real execution after the paused node.

        Args:
            workflow: Workflow dict with ``nodes`` and ``edges``.
            approval_id: ID of the approval record that paused this run.
            approved: Whether the approval was approved (continue) or
                rejected (stop and report ``rejected: True``).

        Yields:
            JSON-encoded (NDJSON) strings, including ``error`` (if the
            approval record or its resume context is missing), ``wf_resumed``,
            and a final ``wf_complete``.
        """
        appr = _ss_exec(
            'SELECT * FROM hub_approvals WHERE approval_id=?',
            (approval_id,), fetchone=True)
        if not appr:
            yield json.dumps({'type': 'error', 'message': 'Approval record not found'}) + '\n'
            return

        appr = _hub_bp._fix_row(appr)
        ctx_data   = json.loads(appr.get('context_json') or '{}')
        resume_data = ctx_data.get('resume', {})

        if not resume_data:
            yield json.dumps({'type': 'error',
                              'message': 'No resume context — approval was created with an older version'}) + '\n'
            return

        # Restore execution context
        saved   = resume_data.get('context', {})
        context = {
            'last_output':       saved.get('last_output', ''),
            'outputs':           saved.get('outputs', {}),
            'vars':              saved.get('vars', {}),
            'condition_results': {k: bool(v) for k, v in saved.get('condition_results', {}).items()},
        }

        nodes     = workflow.get('nodes', [])
        edges     = workflow.get('edges', [])
        order     = resume_data.get('execution_order') or self._engine._topo_sort(nodes, edges)
        paused_at = resume_data.get('paused_at_node', '')

        wf_name = workflow.get('name', '')
        yield json.dumps({'type': 'wf_resumed', 'workflow': wf_name,
                          'approved': approved, 'approval_id': approval_id}) + '\n'

        if not approved:
            yield json.dumps({
                'type':         'wf_complete',
                'final_output': context.get('last_output', ''),
                'vars':         {k: str(v)[:500] for k, v in context.get('vars', {}).items()},
                'rejected':     True,
            }) + '\n'
            return

        # Pass the FULL order + paused_at to _execute_nodes so it can dry-run
        # the pre-pause nodes and correctly rebuild the `skipped` cascade set.
        # Nodes that were already executed are harmlessly re-checked (condition
        # results are in context so their skip/run outcome is deterministic).
        has_nodes_after = paused_at in order and order.index(paused_at) < len(order) - 1
        if not has_nodes_after:
            yield json.dumps({
                'type':         'wf_complete',
                'final_output': context.get('last_output', ''),
                'vars':         {k: str(v)[:500] for k, v in context.get('vars', {}).items()},
            }) + '\n'
            return

        yield from self._execute_nodes(nodes, edges, order, context,
                                       resume_after=paused_at)

    # ── Parallel execution ────────────────────────────────────────────────────

    def _run_parallel(self, nodes, input_data):
        """Run every agent node against the same input and combine their results.

        Note: like the base ``WorkflowEngine._run_parallel``, agent nodes are
        run synchronously in sequence, not concurrently — "parallel" refers
        to the workflow's fan-out structure rather than actual concurrency.
        Non-agent node types and approval pausing are not supported here.

        Args:
            nodes: All workflow nodes; only ``type == 'agent'`` nodes run.
            input_data: Input value passed identically to every agent node.

        Yields:
            JSON-encoded (NDJSON) ``node_complete`` events per agent node,
            then one ``wf_complete`` with all results concatenated.
        """
        results = []
        for node in nodes:
            if node.get('type') == 'agent' and node.get('agent_id'):
                agent = self._get_agent(node['agent_id'])
                if agent:
                    if _hub_bp._APPROVAL_ENABLED:
                        _hub_bp._ensure_approval_tool()
                    tool_configs = _hub_bp._extract_tool_configs(agent.get('tools', []))
                    hub_ctx = {
                        'user':           self._user,
                        'agent_id':       node['agent_id'],
                        'agent_name':     agent.get('name', ''),
                        'convo_id':       self._run_id,
                        'agent_env_vars': agent.get('env_vars', {}),
                    }
                    orch = HubOrchestrator(self._engine.api_key, hub_ctx, tool_configs)
                    result_content = ''
                    for chunk in orch.run(input_data, agent):
                        try:
                            d = json.loads(chunk.strip())
                            if d.get('type') == 'final':
                                result_content = d.get('content', '')
                        except Exception:
                            pass
                    results.append({'node_id': node['id'], 'result': result_content})
                    yield json.dumps({'type': 'node_complete', 'node_id': node['id'],
                                      'result': result_content}) + '\n'
        combined = '\n\n'.join([f"**{r['node_id']}**: {r['result']}" for r in results])
        yield json.dumps({'type': 'wf_complete', 'final_output': combined}) + '\n'
